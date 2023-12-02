from openpyxl import load_workbook

file_path = '30112023 Arquitectura Curricular - Proyecto de Formación.xlsx'

wb = load_workbook(file_path)
print(wb.sheetnames)