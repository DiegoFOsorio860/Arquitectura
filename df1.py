from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

import pandas as pd


file_path = '30112023 Arquitectura Curricular - Proyecto de Formación.xlsx'
# Cargar el archivo Excel usando pandas para mantener las fórmulas intactas
# Cargar el archivo Excel usando pandas para mantener el formato y las fórmulas intactas
new_file_path = 'archivo_con_valores.xlsx'

wb = load_workbook(file_path)
# Seleccionar la hoja de cálculo "TIME"
sheet_name = 'TIME'
ws = wb[sheet_name]

# Crear un nuevo libro de trabajo
new_wb = pd.ExcelWriter('archivo_con_valores.xlsx', engine='openpyxl')
new_wb.book = wb

# Seleccionar la hoja de cálculo del nuevo libro
new_ws = new_wb.book[sheet_name]

# Iterar sobre cada celda y copiar el valor o el resultado de la fórmula
for row in ws.iter_rows(values_only=True):
    new_row = []
    for cell_value in row:
        if isinstance(cell_value, str) and cell_value.startswith('='):
            try:
                new_row.append(ws.cell(1, 1).value)  # Agregar aquí la lógica para calcular el resultado de la fórmula
            except:
                new_row.append(None)  # En caso de error al calcular la fórmula
        else:
            new_row.append(cell_value)
    new_ws.append(new_row)

# Guardar el nuevo archivo
new_wb.save()
new_wb.close()